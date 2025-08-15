# graph_client.py
import os
import io
import time
import json
import urllib.parse
import hashlib
from datetime import datetime
from typing import Any, Dict, List, Optional, Iterable

import requests
import msal
from dotenv import load_dotenv

load_dotenv()

TENANT_ID = os.getenv("TENANT_ID", "common")
CLIENT_ID = os.getenv("CLIENT_ID", "")

# -------------------- SCOPES --------------------
# Permite definir en .env: GRAPH_SCOPES="Files.ReadWrite.All,offline_access,openid,profile,User.Read"
_ENV_SCOPES = os.getenv("GRAPH_SCOPES", "").strip()


def _to_scope_list(values: Optional[Iterable]) -> List[str]:
    """
    Convierte cualquier cosa (str CSV, lista, tupla, set, frozenset) a list[str].
    Elimina vacíos y recorta espacios.
    """
    if not values:
        return []
    if isinstance(values, str):
        parts = [p.strip() for p in values.split(",")]
        return [p for p in parts if p]
    if isinstance(values, (set, frozenset, tuple, list)):
        return [str(p).strip() for p in list(values) if str(p).strip()]
    # iterable genérico
    try:
        return [str(p).strip() for p in list(values) if str(p).strip()]
    except Exception:
        return [str(values).strip()]

# Scopes por defecto (válidos para OneDrive + login interactivo)
_DEFAULT_SCOPES = [
    "Files.ReadWrite.All",
    "User.Read",
    "offline_access",
    "openid",
    "profile",
]

SCOPES: List[str] = _to_scope_list(_ENV_SCOPES) or list(_DEFAULT_SCOPES)

EXCEL_PATH = os.getenv("EXCEL_PATH", "/me/drive/root:/Documentos/resultados/central_solicitudes.xlsx")
COMPROBANTES_ROOT = os.getenv("COMPROBANTES_ROOT", "/me/drive/root:/Comprobantes")

TOKEN_CACHE_FILE = os.getenv("TOKEN_CACHE_FILE", "msal_token_cache.json")

WORKSHEET_NAME = os.getenv("WORKSHEET_NAME", "Hoja1")
TABLE_NAME = os.getenv("TABLE_NAME", "Solicitudes")

CLIENTS_WORKSHEET = os.getenv("CLIENTS_WORKSHEET", "Clientes")
CLIENTS_TABLE = os.getenv("CLIENTS_TABLE", "Clientes")

DEPS_WORKSHEET = os.getenv("DEPS_WORKSHEET", "Depositos")
DEPS_TABLE = os.getenv("DEPS_TABLE", "Deps")

GRAPH_API = "https://graph.microsoft.com/v1.0"


# -------------------- AUTH --------------------
def _load_cache() -> msal.SerializableTokenCache:
    cache = msal.SerializableTokenCache()
    if os.path.exists(TOKEN_CACHE_FILE):
        with open(TOKEN_CACHE_FILE, "r", encoding="utf-8") as f:
            cache.deserialize(f.read())
    return cache

def _save_cache(cache: msal.SerializableTokenCache):
    if cache.has_state_changed:
        with open(TOKEN_CACHE_FILE, "w", encoding="utf-8") as f:
            f.write(cache.serialize())

def get_token() -> str:
    if not CLIENT_ID:
        raise RuntimeError("Falta CLIENT_ID en .env")

    scopes = list(SCOPES)  # ← asegura lista

    cache = _load_cache()
    app = msal.PublicClientApplication(
        client_id=CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{TENANT_ID}",
        token_cache=cache,
    )
    # 1) silencioso si hay cuenta
    accounts = app.get_accounts()
    if accounts:
        res = app.acquire_token_silent(scopes, account=accounts[0])
        if res and "access_token" in res:
            _save_cache(cache)
            return res["access_token"]

    # 2) device code flow
    flow = app.initiate_device_flow(scopes=scopes)  # ← lista, no set
    if "user_code" not in flow:
        raise RuntimeError(f"No se pudo iniciar device code flow: {flow}")
    print("\n== Autenticación requerida ==\n" + flow["message"])
    res = app.acquire_token_by_device_flow(flow)
    if "access_token" not in res:
        raise RuntimeError(f"Fallo autenticación: {res.get('error')} - {res.get('error_description')}")
    _save_cache(cache)
    return res["access_token"]

def _h(token: str, content_json=True) -> Dict[str, str]:
    h = {"Authorization": f"Bearer {token}"}
    if content_json:
        h["Content-Type"] = "application/json"
    return h


# -------------------- OneDrive helpers --------------------
def _get_driveitem_by_path(token: str, path: str) -> Dict[str, Any]:
    url = f"{GRAPH_API}{path}"
    r = requests.get(url, headers=_h(token), timeout=60)
    if r.status_code != 200:
        raise FileNotFoundError(f"No se encontró recurso {path}: {r.status_code} {r.text[:200]}")
    return r.json()

def _get_item_json(token: str, item_id: str) -> Dict[str, Any]:
    r = requests.get(f"{GRAPH_API}/me/drive/items/{item_id}", headers=_h(token), timeout=60)
    if r.status_code != 200:
        raise RuntimeError(f"No se pudo recuperar item: {r.status_code} {r.text[:200]}")
    return r.json()

def _ensure_folder(token: str, parent_item_id: str, name: str) -> str:
    url = f"{GRAPH_API}/me/drive/items/{parent_item_id}/children?$select=id,name,folder"
    r = requests.get(url, headers=_h(token), timeout=60)
    if r.status_code != 200:
        raise RuntimeError(f"No se pudieron listar hijos: {r.status_code} {r.text[:200]}")
    for it in r.json().get("value", []):
        if it.get("folder") and it["name"] == name:
            return it["id"]
    body = {"name": name, "folder": {}, "@microsoft.graph.conflictBehavior": "rename"}
    rc = requests.post(f"{GRAPH_API}/me/drive/items/{parent_item_id}/children", headers=_h(token), json=body, timeout=60)
    if rc.status_code not in (200, 201):
        raise RuntimeError(f"No se pudo crear carpeta '{name}': {rc.status_code} {rc.text[:200]}")
    return rc.json()["id"]

def _ensure_path_chain(token: str, root_path: str, chain: List[str]) -> Dict[str, Any]:
    root_item = _get_driveitem_by_path(token, root_path)
    current_id = root_item["id"]
    for part in chain:
        current_id = _ensure_folder(token, current_id, part)
    return _get_item_json(token, current_id)


# -------------------- Upload (simple + sesión) --------------------
def upload_large_with_session(token: str, folder_item_id: str, final_name: str, fileobj, chunk_size=8*1024*1024):
    create_url = f"{GRAPH_API}/me/drive/items/{folder_item_id}:/{urllib.parse.quote(final_name)}:/createUploadSession"
    r = requests.post(create_url, headers=_h(token), json={}, timeout=60)
    r.raise_for_status()
    upload_url = r.json()["uploadUrl"]

    fileobj.seek(0, io.SEEK_END)
    total = fileobj.tell()
    fileobj.seek(0)
    start = 0
    attempt = 0

    while start < total:
        end = min(start + chunk_size, total) - 1
        fileobj.seek(start)
        chunk = fileobj.read(end - start + 1)
        headers = {
            "Content-Length": str(len(chunk)),
            "Content-Range": f"bytes {start}-{end}/{total}"
        }
        resp = requests.put(upload_url, headers=headers, data=chunk, timeout=300)
        if resp.status_code in (200, 201):   # terminado
            return resp.json()
        if resp.status_code == 202:          # continuar
            start = end + 1
            attempt = 0
            continue
        if resp.status_code in (429, 503):   # throttle
            wait = float(resp.headers.get("Retry-After", "2"))
            time.sleep(wait)
            attempt += 1
            if attempt > 6:
                raise RuntimeError(f"Throttle persistente: {resp.status_code} {resp.text[:200]}")
            continue
        resp.raise_for_status()

def upload_file_stream(
    token: str,
    file_stream,
    original_filename: str,
    cliente_nombre: str,
    fecha: Optional[datetime] = None,
    rename_safe: bool = True,
) -> Dict[str, Any]:
    fecha = fecha or datetime.now()
    chain = [
        (cliente_nombre or "SIN_CLIENTE").strip() or "SIN_CLIENTE",
        f"{fecha:%Y}", f"{fecha:%m}", f"{fecha:%d}",
    ]
    dest_folder = _ensure_path_chain(token, COMPROBANTES_ROOT, chain)
    folder_id = dest_folder["id"]

    name, ext = os.path.splitext(original_filename)
    safe = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in name).strip("_") or "comprobante"
    h = hashlib.md5((original_filename + str(time.time())).encode("utf-8")).hexdigest()[:8]
    final_name = f"{safe}_{fecha:%Y%m%d_%H%M%S}_{h}{ext.lower()}" if rename_safe else original_filename

    # medir tamaño
    file_stream.seek(0, io.SEEK_END)
    size = file_stream.tell()
    file_stream.seek(0)

    if size < 4 * 1024 * 1024:
        put_url = f"{GRAPH_API}/me/drive/items/{folder_id}:/{urllib.parse.quote(final_name)}:/content"
        r = requests.put(put_url, headers={"Authorization": f"Bearer {token}"}, data=file_stream, timeout=300)
        if r.status_code not in (200, 201):
            raise RuntimeError(f"Error subiendo (PUT simple): {r.status_code} {r.text[:200]}")
        return r.json()
    else:
        return upload_large_with_session(token, folder_id, final_name, file_stream)


# -------------------- Excel base local (si no existe) --------------------
def _generate_local_base_excel(path: str):
    from openpyxl import Workbook
    wb = Workbook()

    # Hoja de detalle
    ws = wb.active
    ws.title = WORKSHEET_NAME
    ws.append([
        "fecha_iso","banco","forma_pago","producto","importe","tipo_bbva",
        "folio","autorizacion","folio_movimiento","numero_usuario","requiere_factura",
        "observaciones","cliente_id","cliente_nombre","comprobante_url","origen"
    ])

    # Hoja de resumen Deps
    ws2 = wb.create_sheet(DEPS_WORKSHEET)
    ws2.append(["Fecha banco","Cliente","Monto","Movimiento","Ficha","Realizó","Observaciones","Factura"])

    # Hoja de Clientes
    ws3 = wb.create_sheet(CLIENTS_WORKSHEET)
    ws3.append(["cliente_id","cliente_nombre","rfc","razon_social","cfdi","uso_cfdi","direccion","contacto","email","numero_usuario"])

    wb.save(path)

# -------------------- Excel (tablas) --------------------
def _ensure_excel_exists(token: str, template_local: Optional[str] = None) -> Dict[str, Any]:
    # Si existe:
    try:
        return _get_driveitem_by_path(token, EXCEL_PATH)
    except FileNotFoundError:
        pass
    # Si no existe: crear local y subir
    if template_local:
        if not os.path.exists(template_local):
            os.makedirs(os.path.dirname(template_local) or ".", exist_ok=True)
            _generate_local_base_excel(template_local)
        put_url = f"{GRAPH_API}{EXCEL_PATH}:/content"
        with open(template_local, "rb") as f:
            data = f.read()
        r = requests.put(put_url, headers={"Authorization": f"Bearer {token}"}, data=data, timeout=300)
        if r.status_code not in (200, 201):
            raise RuntimeError(f"No se pudo subir el Excel base: {r.status_code} {r.text[:200]}")
        return r.json()
    raise FileNotFoundError("No existe el Excel central en OneDrive y no se proporcionó template_local.")

def _ensure_table_exists(token: str, file_id: str, worksheet_name: str, table_name: str) -> str:
    base = f"{GRAPH_API}/me/drive/items/{file_id}/workbook/worksheets('{urllib.parse.quote(worksheet_name)}')"
    # listar tablas
    r = requests.get(f"{base}/tables", headers=_h(token), timeout=60)
    if r.status_code != 200:
        raise RuntimeError(f"No se pudieron listar tablas: {r.status_code} {r.text[:200]}")
    for t in r.json().get("value", []):
        if t.get("name") == table_name:
            return t["id"]

    # crear tabla sobre usedRange (asume encabezados en fila 1)
    ru = requests.get(f"{base}/usedRange(valuesOnly=true)", headers=_h(token), timeout=60)
    if ru.status_code != 200:
        raise RuntimeError(f"No se pudo obtener usedRange: {ru.status_code} {ru.text[:200]}")
    address = ru.json().get("address")  # p.ej. Hoja1!A1:P1
    body = {"address": address, "hasHeaders": True}
    rc = requests.post(f"{base}/tables", headers=_h(token), json=body, timeout=60)
    if rc.status_code not in (200, 201):
        raise RuntimeError(f"No se pudo crear la tabla: {rc.status_code} {rc.text[:200]}")
    tid = rc.json()["id"]
    # renombrar
    rn = requests.patch(
        f"{GRAPH_API}/me/drive/items/{file_id}/workbook/tables/{tid}",
        headers=_h(token),
        json={"name": table_name},
        timeout=60,
    )
    if rn.status_code not in (200, 204):
        raise RuntimeError(f"Tabla creada pero no se pudo renombrar: {rn.status_code} {rn.text[:200]}")
    return tid

def append_rows(token: str, file_id: str, table_id: str, rows: List[List[Any]]):
    url = f"{GRAPH_API}/me/drive/items/{file_id}/workbook/tables/{table_id}/rows/add"
    for attempt in range(3):
        r = requests.post(url, headers=_h(token), json={"values": rows}, timeout=60)
        if r.status_code in (200, 201):
            return
        if r.status_code in (423, 429, 503):
            time.sleep(float(r.headers.get("Retry-After", "2")))
            continue
        raise RuntimeError(f"No se pudieron agregar filas: {r.status_code} {r.text[:200]}")
    raise RuntimeError(f"No se pudieron agregar filas tras reintentos")

def map_payload_to_row(payload: Dict[str, Any]) -> List[Any]:
    return [
        payload.get("fecha_iso") or "",
        payload.get("banco") or "",
        payload.get("forma_pago") or "",
        payload.get("producto") or "",
        payload.get("importe") or "",
        payload.get("tipo_bbva") or "",
        payload.get("folio") or "",
        payload.get("autorizacion") or "",
        payload.get("folio_movimiento") or "",
        payload.get("numero_usuario") or "",
        "Sí" if payload.get("requiere_factura") else "No",
        payload.get("observaciones") or "",
        payload.get("cliente_id") or "",
        payload.get("cliente_nombre") or "",
        payload.get("comprobante_url") or "",
        payload.get("origen") or "formulario",
    ]

def ensure_excel_and_table(token: str, template_local: Optional[str] = None):
    wb_item = _ensure_excel_exists(token, template_local=template_local)
    table_id = _ensure_table_exists(token, wb_item["id"], WORKSHEET_NAME, TABLE_NAME)
    return wb_item, table_id


# -------------------- RESUMEN “Deps” --------------------
import locale
try:
    locale.setlocale(locale.LC_TIME, "es_MX.UTF-8")
except Exception:
    try:
        locale.setlocale(locale.LC_TIME, "es_ES.UTF-8")
    except Exception:
        pass

def _fmt_fecha_dd_mmm(fecha_iso: str) -> str:
    try:
        dt = datetime.fromisoformat(fecha_iso)
        return dt.strftime("%d-%b").lower().replace(".", "")
    except Exception:
        return fecha_iso or ""

def map_payload_to_deps_row(payload: Dict[str, Any]) -> List[Any]:
    # Fecha banco
    fecha_banco = _fmt_fecha_dd_mmm(payload.get("fecha_iso") or "")
    # Cliente (usa cliente_id; si no hay, cliente_nombre)
    cliente = payload.get("cliente_id") or payload.get("cliente_nombre") or ""
    # Monto
    importe = str(payload.get("importe") or "").strip()
    # Movimiento
    banco = (payload.get("banco") or "").upper()
    tipo_bbva = (payload.get("tipo_bbva") or "").lower()
    folio = (payload.get("folio") or "").strip()
    aut = (payload.get("autorizacion") or "").strip()
    folmov = (payload.get("folio_movimiento") or "").strip()
    producto = (payload.get("producto") or "").lower()

    if "tae" in producto:
        movimiento = ""  # si es TAE, vacío
    elif banco == "BBVA" and tipo_bbva == "practicaja" and (folio or aut):
        movimiento = f"Folio {folio}" + (f"/ Aut {aut}" if aut else "")
    elif banco == "BBVA" and tipo_bbva == "ventanilla" and folmov:
        movimiento = f"Folio/Mov {folmov}"
    elif folio:
        movimiento = f"Ref {folio}"
    else:
        movimiento = "-"

    # Ficha
    forma_pago = (payload.get("forma_pago") or "").lower()
    ficha = "s c" if forma_pago in ("depósito", "deposito") else (
        "transfer" if forma_pago == "transferencia" else ""
    )

    # Realizó
    realizo = payload.get("realizo") or ""

    # Observaciones
    extra = (payload.get("observaciones") or "").strip()
    if "tae" in producto:
        observaciones = extra  # para TAE se deja tal cual (suele quedar vacío)
    else:
        base_obs = ".p/pago DEP PS"
        observaciones = base_obs + (f"  {extra}" if extra else "")

    # Factura  ✅ (corregido)
    factura = "Sí" if payload.get("requiere_factura") else "No"

    return [fecha_banco, cliente, importe, movimiento, ficha, realizo, observaciones, factura]


def ensure_excel_and_deps_table(token: str):
    wb_item = _ensure_excel_exists(token)
    table_id = _ensure_table_exists(token, wb_item["id"], DEPS_WORKSHEET, DEPS_TABLE)
    return wb_item, table_id

def add_deps_row(token: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    wb_item, table_id = ensure_excel_and_deps_table(token)
    row = map_payload_to_deps_row(payload)
    append_rows(token, wb_item["id"], table_id, [row])
    return {"ok": True, "excel": wb_item.get("name"), "tabla": DEPS_TABLE}

def add_solicitud_row(token: str, payload: Dict[str, Any], template_local: Optional[str] = None):
    """
    Asegura el Excel + tabla de detalle y agrega una fila.
    Si el Excel no existe en OneDrive, sube un template local generado automáticamente.
    """
    wb_item, table_id = ensure_excel_and_table(token, template_local=template_local)
    row = map_payload_to_row(payload)
    append_rows(token, wb_item["id"], table_id, [row])
    return {"ok": True, "excel": wb_item.get("name"), "tabla": TABLE_NAME}


# -------------------- Clientes (cards por número de usuario) --------------------
def get_clientes_por_usuario(token: str, numero_usuario: str) -> List[Dict[str, Any]]:
    wb = _get_driveitem_by_path(token, EXCEL_PATH)
    ws_list = requests.get(f"{GRAPH_API}/me/drive/items/{wb['id']}/workbook/worksheets", headers=_h(token), timeout=60)
    if ws_list.status_code != 200:
        return []

    ws_id = None
    for ws in ws_list.json().get("value", []):
        if ws.get("name") == CLIENTS_WORKSHEET:
            ws_id = ws["id"]
            break
    if not ws_id:
        return []

    tabs = requests.get(
        f"{GRAPH_API}/me/drive/items/{wb['id']}/workbook/worksheets('{urllib.parse.quote(CLIENTS_WORKSHEET)}')/tables",
        headers=_h(token), timeout=60)
    if tabs.status_code != 200:
        return []
    table_id = None
    for t in tabs.json().get("value", []):
        if t.get("name") == CLIENTS_TABLE:
            table_id = t["id"]
            break
    if not table_id:
        return []

    rng = requests.get(f"{GRAPH_API}/me/drive/items/{wb['id']}/workbook/tables/{table_id}/range", headers=_h(token), timeout=60)
    if rng.status_code != 200:
        return []
    values = rng.json().get("values", [])
    if not values:
        return []
    headers = [str(h or "").strip() for h in values[0]]
    rows = values[1:]

    idx_num = headers.index("numero_usuario") if "numero_usuario" in headers else None
    idx_id = headers.index("cliente_id") if "cliente_id" in headers else None
    idx_nombre = headers.index("cliente_nombre") if "cliente_nombre" in headers else None
    idx_rfc = headers.index("rfc") if "rfc" in headers else None

    found = []
    for r in rows:
        r = list(r) + [""] * (len(headers) - len(r))
        hit = False
        if idx_num is not None and str(r[idx_num]).strip() == str(numero_usuario).strip():
            hit = True
        elif idx_id is not None and str(r[idx_id]).strip() == str(numero_usuario).strip():
            hit = True
        else:
            if any(str(c).strip() == str(numero_usuario).strip() for c in r):
                hit = True
        if hit:
            found.append({
                "cliente_id": r[idx_id] if idx_id is not None else "",
                "cliente_nombre": r[idx_nombre] if idx_nombre is not None else "",
                "rfc": r[idx_rfc] if idx_rfc is not None else "",
                "raw": dict(zip(headers, r))
            })
    return found
